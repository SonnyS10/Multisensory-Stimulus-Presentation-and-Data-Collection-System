"""
Stimulus Order Management Frame
This module provides a GUI interface for managing the order of stimulus presentation
for each test. Users can view and rearrange the order of images through drag-and-drop.
"""

import os
import csv
import re
import time
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox, 
    QListWidget, QListWidgetItem, QFrame, QMessageBox, QSizePolicy, QFileDialog,
    QCheckBox, QLineEdit, QDialog, QFormLayout, QDialogButtonBox, QTreeWidget, QTreeWidgetItem
)
from PyQt5.QtGui import QFont, QPixmap, QIcon
from PyQt5.QtCore import Qt, QSize
from eeg_stimulus_project.assets.asset_handler import Display
import openpyxl

class StimulusOrderFrame(QWidget):
    """
    A frame that allows users to view and rearrange the order of stimulus presentation
    for each test using drag-and-drop functionality.
    """
    
    def __init__(self, parent=None, alcohol_folder=None, non_alcohol_folder=None):
        super().__init__(parent)
        self.parent = parent
        self.alcohol_folder = alcohol_folder
        self.non_alcohol_folder = non_alcohol_folder
        self.current_test_name = None
        self.custom_orders = {}  # Store applied custom orders for each test (what's actually used for tests)
        self.working_orders = {}  # Store working orders for each test (what user is currently editing)
        self.original_assets = {}  # Store original asset order
        self.scent_numbers = {}  # {asset: scent_number}
        
        # Add randomization and repetitions settings
        self.randomize_cues = False
        self.seed = None
        self.stimulus_repetitions = {}  # Store repetitions here
        
        # Setup UI first
        self.setup_ui()
        
        # Load current assets
        self.load_current_assets()
        
        # Set default test
        if self.test_selector.count() > 0:
            self.test_selector.setCurrentIndex(0)
            self.on_test_selected()
    
    def setup_ui(self):
        """Setup the user interface components."""
        #print("Setting up Stimulus Order Frame UI...")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Title
        title = QLabel("Stimulus Order Management")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Test selection
        test_layout = QHBoxLayout()
        test_label = QLabel("Select Test:")
        test_label.setFont(QFont("Segoe UI", 12))
        test_layout.addWidget(test_label)
        
        self.test_selector = QComboBox()
        self.test_selector.setFont(QFont("Segoe UI", 11))
        self.test_selector.setMinimumWidth(400)
        
        # Populate test selector with available tests
        test_names = [
            'Unisensory Neutral Visual',
            'Unisensory Alcohol Visual',
            'Multisensory Neutral Visual & Olfactory',
            'Multisensory Alcohol Visual & Olfactory',
            'Multisensory Neutral Visual, Tactile & Olfactory',
            'Multisensory Alcohol Visual, Tactile & Olfactory',
            'Stroop Multisensory Alcohol (Visual & Tactile)',
            'Stroop Multisensory Neutral (Visual & Tactile)',
            'Stroop Multisensory Alcohol (Visual & Olfactory)',
            'Stroop Multisensory Neutral (Visual & Olfactory)'
        ]
        
        for test_name in test_names:
            self.test_selector.addItem(test_name)
        
        test_layout.addWidget(self.test_selector)
        test_layout.addStretch()

        # --- Go to Selected Test Button (top right) ---
        self.goto_test_btn = QPushButton("Go to Selected Test")
        self.goto_test_btn.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.goto_test_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff7043;
                color: white;
                border-radius: 8px;
                padding: 12px 32px;
                font-size: 16px;
                min-width: 180px;
                min-height: 44px;
            }
            QPushButton:hover {
                background-color: #d84315;
            }
        """)
        self.goto_test_btn.clicked.connect(self.goto_selected_test)
        test_layout.addWidget(self.goto_test_btn)

        layout.addLayout(test_layout)
        
        # Instructions
        instructions = QLabel(
            "Drag and drop images to rearrange their presentation order. "
            "Changes are made to your working order. Click 'Apply Custom Order' to save changes for actual test use."
        )
        instructions.setFont(QFont("Segoe UI", 14))
        instructions.setTextFormat(Qt.RichText)
        instructions.setWordWrap(True)
        instructions.setStyleSheet("color: #000000; margin: 10px 0;")
        layout.addWidget(instructions)
        
        # Image list
        list_frame = QFrame()
        list_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        list_layout = QVBoxLayout(list_frame)
        
        list_label = QLabel("Working Order (drag to rearrange):")
        list_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        list_layout.addWidget(list_label)
        
        self.image_list = QListWidget()
        self.image_list.setDragDropMode(QListWidget.InternalMove)
        self.image_list.setDefaultDropAction(Qt.MoveAction)
        self.image_list.setStyleSheet("""
            QListWidget {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                margin: 2px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: #f9f9f9;
            }
            QListWidget::item:selected {
                background-color: #007bff;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #e9ecef;
            }
        """)
        self.image_list.setMinimumHeight(300)
        list_layout.addWidget(self.image_list)
        self.image_list.model().rowsMoved.connect(self.on_rows_moved)
        
        layout.addWidget(list_frame)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        reset_button = QPushButton("Reset Working Order")
        reset_button.setFont(QFont("Segoe UI", 11))
        reset_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        reset_button.clicked.connect(self.reset_to_original)
        button_layout.addWidget(reset_button)
        
        button_layout.addStretch()
        
        self.apply_button = QPushButton("Apply Custom Order")
        self.apply_button.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.apply_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.apply_button.clicked.connect(self.apply_custom_order)
        button_layout.addWidget(self.apply_button)
        
        import_button = QPushButton("Import Order from CSV")
        import_button.setFont(QFont("Segoe UI", 11))
        import_button.setStyleSheet("""
            QPushButton {
                background-color: #ffb300;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 160px;
            }
            QPushButton:hover {
                background-color: #ffa000;
            }
        """)
        import_button.clicked.connect(self.import_order_from_csv)
        button_layout.insertWidget(button_layout.count() // 2, import_button)
        
        # --- Add Assign Scent Numbers Button ---
        self.assign_scent_btn = QPushButton("Assign Scent Numbers")
        self.assign_scent_btn.setFont(QFont("Segoe UI", 11))
        self.assign_scent_btn.setStyleSheet("""
            QPushButton {
                background-color: #8e24aa;
                color: white;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 160px;
            }
            QPushButton:hover {
                background-color: #6d1b7b;
            }
        """)
        self.assign_scent_btn.clicked.connect(self.open_scent_assignment_dialog)
        button_layout.addWidget(self.assign_scent_btn)
        
        layout.addLayout(button_layout)

        # --- Available Assets Section ---
        assets_frame = QFrame()
        assets_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
                margin-top: 12px;
            }
        """)
        assets_layout = QVBoxLayout(assets_frame)
        assets_layout.setSpacing(8)

        assets_label = QLabel("Available Assets (Select then hit the Add Selected Asset Button to Add):")
        assets_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        assets_label.setAlignment(Qt.AlignLeft)
        assets_layout.addWidget(assets_label)

        self.available_assets_tree = QTreeWidget()
        self.available_assets_tree.setHeaderHidden(True)
        self.available_assets_tree.setSelectionMode(QTreeWidget.SingleSelection)
        self.available_assets_tree.setMinimumHeight(120)
        self.available_assets_tree.setStyleSheet("""
            QTreeWidget {
                background-color: white;
                border: 1px solid #bbb;
                border-radius: 4px;
                padding: 5px;
            }
            QTreeWidget::item {
                padding: 6px;
                border-radius: 3px;
            }
            QTreeWidget::item:selected {
                background-color: #007bff;
                color: white;
            }
            QTreeWidget::item:hover {
                background-color: #e9ecef;
            }
            QTreeWidget::branch {
                background-color: white;
            }
            QTreeWidget::branch:has-children:!has-siblings:closed,
            QTreeWidget::branch:closed:has-children:has-siblings {
                border-image: none;
                image: url(none);
            }
            QTreeWidget::branch:open:has-children:!has-siblings,
            QTreeWidget::branch:open:has-children:has-siblings {
                border-image: none;
                image: url(none);
            }
        """)
        self.available_assets_tree.setIndentation(20)
        assets_layout.addWidget(self.available_assets_tree)

        asset_btn_layout = QHBoxLayout()
        self.add_asset_btn = QPushButton("Add Selected Asset")
        self.add_asset_btn.setFont(QFont("Segoe UI", 11))
        self.add_asset_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 18px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.add_asset_btn.clicked.connect(self.add_selected_asset_to_test)
        asset_btn_layout.addWidget(self.add_asset_btn)

        self.delete_asset_btn = QPushButton("Delete Selected Stimulus")
        self.delete_asset_btn.setFont(QFont("Segoe UI", 11))
        self.delete_asset_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 18px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.delete_asset_btn.clicked.connect(self.delete_selected_stimulus_from_test)
        asset_btn_layout.addWidget(self.delete_asset_btn)

        asset_btn_layout.addStretch()
        assets_layout.addLayout(asset_btn_layout)

        layout.addWidget(assets_frame)

        # --- Randomization Section ---
        randomization_frame = QFrame()
        randomization_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
                margin-top: 12px;
            }
        """)
        randomization_layout = QVBoxLayout(randomization_frame)
        randomization_layout.setSpacing(8)

        randomization_label = QLabel("Randomization Settings:")
        randomization_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        randomization_layout.addWidget(randomization_label)

        # Randomizer row
        randomizer_row = QHBoxLayout()
        self.randomize_checkbox = QCheckBox("Randomize Alcohol/Non-Alcohol Cues")
        self.randomize_checkbox.setFont(QFont("Segoe UI", 10))

        # Add repetitions checkbox here
        self.repetition_checkbox = QCheckBox("Specify stimulus repetitions")
        self.repetition_checkbox.setFont(QFont("Segoe UI", 10))
        self.repetition_checkbox.setChecked(False)

        self.seed_label = QLabel("Seed(1-10000):")
        self.seed_label.setFont(QFont("Segoe UI", 10))
        self.seed_input = QLineEdit()
        self.seed_input.setFont(QFont("Segoe UI", 10))
        self.seed_input.setPlaceholderText("Leave blank for random")

        randomizer_row.addWidget(self.randomize_checkbox)
        randomizer_row.addWidget(self.repetition_checkbox)
        randomizer_row.addWidget(self.seed_label)
        randomizer_row.addWidget(self.seed_input)
        randomization_layout.addLayout(randomizer_row)

        # --- Add Randomize Now Button ---
        self.randomize_now_btn = QPushButton("Randomize Now")
        self.randomize_now_btn.setFont(QFont("Segoe UI", 10, QFont.Bold))
        self.randomize_now_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 18px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.randomize_now_btn.clicked.connect(self.on_randomize_now_clicked)
        randomization_layout.addWidget(self.randomize_now_btn)

        layout.addWidget(randomization_frame)
        self.setLayout(layout)

        # Now connect the signal, after all widgets are created!
        self.test_selector.currentTextChanged.connect(self.on_test_selected)

    def load_current_assets(self):
        """Load current assets from the asset handler."""
        #print("Loading current assets...")
        try:
            passive_default_limit = 8
            randomize_cues, seed = self.get_randomization_settings()
            repetitions = self.get_repetitions_settings()
            
            self.original_assets = Display.get_assets(
                alcohol_folder=self.alcohol_folder,
                non_alcohol_folder=self.non_alcohol_folder,
                randomize_cues=randomize_cues,
                seed=seed,
                repetitions=repetitions
            )
            # --- Ensure CravingRatingAsset is last in all passive tests ---
            for test_name, asset_list in self.original_assets.items():
                if test_name and not test_name.lower().startswith("stroop"):
                    asset_list = [a for a in asset_list if not isinstance(a, CravingRatingAsset)]
                    asset_list = asset_list[:passive_default_limit]
                    # Remove any existing craving rating asset
                    original_craving = CravingRatingAsset()
                    original_craving.is_original = True
                    asset_list.append(original_craving)
                    self.original_assets[test_name] = asset_list

            # Gather all unique assets for the available assets list, tagging by origin
            asset_dict = {}
            
            # Get custom folder paths if provided and validate they're not just the project root
            gui_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(gui_dir)  # eeg_stimulus_project
            project_root = os.path.normpath(os.path.abspath(project_root))
            
            # Only use custom folders if they're actually set and not the project root
            alcohol_folder = None
            if self.alcohol_folder and os.path.abspath(self.alcohol_folder) != project_root:
                alcohol_folder = os.path.normpath(os.path.abspath(self.alcohol_folder))
            
            non_alcohol_folder = None
            if self.non_alcohol_folder and os.path.abspath(self.non_alcohol_folder) != project_root:
                non_alcohol_folder = os.path.normpath(os.path.abspath(self.non_alcohol_folder))
            
            # Get default folder paths
            default_alcohol_folder = os.path.join(project_root, 'assets', 'Images', 'Default', 'Alcohol')
            default_neutral_folder = os.path.join(project_root, 'assets', 'Images', 'Default', 'Neutral')
            
            # Normalize paths for comparison
            default_alcohol_folder = os.path.normpath(os.path.abspath(default_alcohol_folder))
            default_neutral_folder = os.path.normpath(os.path.abspath(default_neutral_folder))
            
            print(f"=== Asset Origin Detection ===")
            print(f"Project Root: {project_root}")
            print(f"Default Alcohol Folder: {default_alcohol_folder}")
            print(f"Default Neutral Folder: {default_neutral_folder}")
            print(f"Custom Alcohol Folder: {alcohol_folder}")
            print(f"Custom Neutral Folder: {non_alcohol_folder}")
            print(f"=============================")
            
            for images in self.original_assets.values():
                for img in images:
                    fname = getattr(img, 'filename', None)
                    if fname:
                        fname_abs = os.path.normpath(os.path.abspath(fname))
                        fname_dir = os.path.dirname(fname_abs)

                        # Check paths - be very specific with matching
                        # Check if in alcohol folders (default or custom)
                        is_alcohol = False
                        is_neutral = False
                        
                        # Check default folders (exact match)
                        if fname_dir == default_alcohol_folder:
                            is_alcohol = True
                        elif fname_dir == default_neutral_folder:
                            is_neutral = True
                        # Check custom folders (exact match or subdirectory)
                        elif alcohol_folder and (fname_dir == alcohol_folder or fname_dir.startswith(alcohol_folder + os.sep)):
                            is_alcohol = True
                        elif non_alcohol_folder and (fname_dir == non_alcohol_folder or fname_dir.startswith(non_alcohol_folder + os.sep)):
                            is_neutral = True
                        
                        # Assign origin based on checks
                        if is_alcohol:
                            img.asset_origin = "alcohol"
                        elif is_neutral:
                            img.asset_origin = "neutral"
                        else:
                            # Fallback: check if filename contains alcohol-related keywords
                            fname_lower = fname.lower()
                            if any(keyword in fname_lower for keyword in ['beer', 'stella', 'wine', 'whiskey', 'vodka', 'alcohol', 'liquor']):
                                img.asset_origin = "alcohol"
                            else:
                                img.asset_origin = "unknown"
                        
                        # Debug output
                        print(f"  {os.path.basename(fname)} | Dir: {fname_dir} | Origin: {img.asset_origin}")
                        
                        # Use full normalized path as key to avoid overwriting assets with same filename
                        norm = self.normalize_name(fname_abs)
                        asset_dict[norm] = img
            self.all_asset_objs = list(asset_dict.values())
            self.update_available_assets_list()
        except Exception as e:
            print(f"Error loading assets: {e}")
            self.original_assets = {}
            self.all_asset_objs = []
            self.update_available_assets_list()

    def update_available_assets_list(self):
        """Update the available assets tree widget with folder structure, filtering by test type."""
        self.available_assets_tree.clear()
        if not self.current_test_name:
            return

        test_name = self.current_test_name.lower()
        if "neutral" in test_name and "alcohol" not in test_name:
            cue_type = "neutral"
        elif "alcohol" in test_name:
            cue_type = "alcohol"
        else:
            cue_type = None  # Show all if not specified

        # Organize assets by source
        default_assets = []
        custom_assets = []
        
        # Get paths for comparison
        gui_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(gui_dir)
        default_folder = os.path.normpath(os.path.join(project_root, 'assets', 'Images', 'Default'))
        
        print(f"\n=== Tree Display Debug ===")
        print(f"Default folder path: {default_folder}")
        print(f"Filter type: {cue_type}")
        
        for img in self.all_asset_objs:
            origin = getattr(img, "asset_origin", "not_set")
            # Only show assets from the correct type
            if cue_type and origin != cue_type:
                continue
            
            fname = getattr(img, 'filename', None)
            if fname:
                fname_abs = os.path.normpath(os.path.abspath(fname))
                fname_dir = os.path.dirname(fname_abs)
                # Check if it's from default folder or custom
                # Must check if the file path contains the default folder path
                is_default = fname_abs.startswith(default_folder + os.sep) or fname_dir == default_folder
                print(f"  File: {os.path.basename(fname)} | Path: {fname_abs[:80]}... | Is Default: {is_default}")
                
                if is_default:
                    default_assets.append(img)
                else:
                    custom_assets.append(img)
        
        # Create folder structure
        if default_assets:
            default_folder_item = QTreeWidgetItem(self.available_assets_tree)
            default_folder_item.setText(0, f"📁 Default {cue_type.capitalize() if cue_type else ''} Assets ({len(default_assets)})")
            default_folder_item.setFont(0, QFont("Segoe UI", 11, QFont.Bold))
            default_folder_item.setExpanded(True)
            
            for img in default_assets:
                fname = getattr(img, 'filename', None)
                display_name = os.path.splitext(os.path.basename(fname))[0] if fname else "Image"
                item = QTreeWidgetItem(default_folder_item)
                item.setText(0, display_name)
                item.setData(0, Qt.UserRole, img)
                
                if fname:
                    try:
                        pixmap = QPixmap(fname)
                        if not pixmap.isNull():
                            thumbnail = pixmap.scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            item.setIcon(0, QIcon(thumbnail))
                    except Exception as e:
                        print(f"Error creating thumbnail for {fname}: {e}")
        
        if custom_assets:
            custom_folder_item = QTreeWidgetItem(self.available_assets_tree)
            custom_folder_item.setText(0, f"📁 Custom {cue_type.capitalize() if cue_type else ''} Assets ({len(custom_assets)})")
            custom_folder_item.setFont(0, QFont("Segoe UI", 11, QFont.Bold))
            custom_folder_item.setExpanded(True)
            
            for img in custom_assets:
                fname = getattr(img, 'filename', None)
                display_name = os.path.splitext(os.path.basename(fname))[0] if fname else "Image"
                item = QTreeWidgetItem(custom_folder_item)
                item.setText(0, display_name)
                item.setData(0, Qt.UserRole, img)
                
                if fname:
                    try:
                        pixmap = QPixmap(fname)
                        if not pixmap.isNull():
                            thumbnail = pixmap.scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            item.setIcon(0, QIcon(thumbnail))
                    except Exception as e:
                        print(f"Error creating thumbnail for {fname}: {e}")
        
        # Add craving rating asset
        craving_item = QTreeWidgetItem(self.available_assets_tree)
        craving_item.setText(0, "📊 craving_rating")
        craving_item.setFont(0, QFont("Segoe UI", 10, QFont.Bold))
        craving_item.setData(0, Qt.UserRole, CravingRatingAsset())
        
        # Debug output
        total_assets = len(default_assets) + len(custom_assets)
        print(f"Test: {self.current_test_name}, Type filter: {cue_type}, Assets shown: {total_assets}/{len(self.all_asset_objs)} (Default: {len(default_assets)}, Custom: {len(custom_assets)})")

    def on_test_selected(self):
        """Handle test selection change."""
        self.current_test_name = self.test_selector.currentText()
        # Show/hide the assign scent button based on test type
        if "olfactory" in self.current_test_name.lower():
            self.assign_scent_btn.show()
        else:
            self.assign_scent_btn.hide()
        self.update_image_list()
        self.update_available_assets_list()

    def update_image_list(self):
        """Update the image list widget with current test's working order."""
        if not self.current_test_name or self.current_test_name not in self.original_assets:
            return
        
        self.image_list.clear()
        
        # Always use working order, initialize it if it doesn't exist
        if self.current_test_name not in self.working_orders:
            # Initialize working order with current applied order or original order
            if self.current_test_name in self.custom_orders:
                self.working_orders[self.current_test_name] = self.custom_orders[self.current_test_name].copy()
            else:
                self.working_orders[self.current_test_name] = self.original_assets[self.current_test_name].copy()

        images = self.working_orders[self.current_test_name][:]

        # --- Passive test: only move the original craving rating to the end ---
        if self.is_passive_test():
            orig_idx = next((i for i, img in enumerate(images)
                             if isinstance(img, CravingRatingAsset) and getattr(img, "is_original", False)), None)
            if orig_idx is not None and orig_idx != len(images) - 1:
                orig_craving = images.pop(orig_idx)
                images.append(orig_craving)
                self.working_orders[self.current_test_name] = images

        # Determine if scent information should be shown
        show_scent = "olfactory" in self.current_test_name.lower()
        for i, image in enumerate(images):
            item = QListWidgetItem()
            key = getattr(image, "filename", None)
            scent = self.scent_numbers.get(key, None)
            scent_str = f" [Scent: {scent}]" if (show_scent and scent) else ""
            # Check if the image is a CravingRatingAsset
            if isinstance(image, CravingRatingAsset):
                item.setText(f"{i+1}. craving_rating")
                # Make only the original unselectable and unmovable
                if getattr(image, "is_original", False):
                    item.setFlags(item.flags() & ~Qt.ItemIsSelectable & ~Qt.ItemIsDragEnabled)
            elif hasattr(image, 'filename'):
                filename = os.path.basename(image.filename)
                display_name = os.path.splitext(filename)[0]
                
                # Try to create a thumbnail
                try:
                    pixmap = QPixmap(image.filename)
                    if not pixmap.isNull():
                        # Scale to thumbnail size
                        thumbnail = pixmap.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        item.setIcon(QIcon(thumbnail))
                except Exception as e:
                    print(f"Error creating thumbnail for {filename}: {e}")
                
                item.setText(f"{i+1}. {display_name}{scent_str}")
            else:
                item.setText(f"{i+1}. Image {i+1}{scent_str}")
            
            # Store the original image object in the item data
            item.setData(Qt.UserRole, image)
            self.image_list.addItem(item)
        self.update_apply_button_state()

    def reset_to_original(self):
        """Reset the current test's working order to the original image order."""
        if not self.current_test_name:
            return
        
        reply = QMessageBox.question(
            self,
            "Reset Order",
            f"Are you sure you want to reset the working order for '{self.current_test_name}' to the original order?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Reset working order to original order
            self.working_orders[self.current_test_name] = self.original_assets[self.current_test_name].copy()
            
            # Update the display
            self.update_image_list()
            self.update_apply_button_state()
            
    def apply_custom_order(self):
        """Apply the current working order as the custom order for the selected test."""
        if not self.current_test_name:
            return
        
        # Ensure working order exists and is current with UI
        self.sync_working_order_with_ui()

        if not self.validate_passive_unique_stimulus_limit():
            return
        
        # Copy working order to applied custom order
        self.custom_orders[self.current_test_name] = self.working_orders[self.current_test_name].copy()
        
        # Update the parent's asset handler to use custom order
        if hasattr(self.parent, 'update_custom_orders'):
            self.parent.update_custom_orders(self.custom_orders)
        
        QMessageBox.information(
            self,
            "Order Applied",
            f"Custom order applied for '{self.current_test_name}'. "
            f"This order will be used when running the test."
        )
        self.update_apply_button_state()
    
    def sync_working_order_with_ui(self):
        """Synchronize the working order with the current UI state."""
        if not self.current_test_name:
            return
        
        # Get current order from the list widget
        current_order = []
        for i in range(self.image_list.count()):
            item = self.image_list.item(i)
            image = item.data(Qt.UserRole)
            current_order.append(image)
        
        # Update working order
        self.working_orders[self.current_test_name] = current_order
    
    def on_rows_moved(self, *args):
        """Handle when rows are moved via drag and drop."""
        self.sync_working_order_with_ui()
        # Passive: only force the original craving rating to the end
        if self.is_passive_test():
            working = self.working_orders[self.current_test_name]
            # Find the original craving rating asset
            orig_idx = next((i for i, asset in enumerate(working)
                            if isinstance(asset, CravingRatingAsset) and getattr(asset, "is_original", False)), None)
            if orig_idx is not None and orig_idx != len(working) - 1:
                orig_craving = working.pop(orig_idx)
                working.append(orig_craving)
        self.update_apply_button_state()
        self.update_image_list()
    
    def add_selected_asset_to_test(self):
        """Add the selected asset from the available tree to the current test's working order."""
        if not self.current_test_name:
            return
        selected_items = self.available_assets_tree.selectedItems()
        if not selected_items:
            return
        
        # Get the selected item - could be a folder or an asset
        selected_item = selected_items[0]
        img = selected_item.data(0, Qt.UserRole)
        
        # If it's a folder (no data), don't add
        if img is None:
            QMessageBox.information(self, "Select Asset", "Please select an asset, not a folder.")
            return
        
        # Ensure working order exists
        if self.current_test_name not in self.working_orders:
            self.working_orders[self.current_test_name] = self.original_assets[self.current_test_name].copy()
        
        if self.is_passive_test():
            working = self.working_orders[self.current_test_name]
            craving_idx = next((i for i, asset in enumerate(working) if isinstance(asset, CravingRatingAsset)), None)
            if craving_idx is not None:
                working.insert(craving_idx, img)
            else:
                working.append(img)

            if not self.validate_passive_unique_stimulus_limit():
                # Revert this add if it violates passive unique-stimulus constraints.
                if img in working:
                    working.remove(img)
                self.update_image_list()
                self.update_apply_button_state()
                return
        else:
            self.working_orders[self.current_test_name].append(img)
        self.update_image_list()
        self.update_apply_button_state()

    def delete_selected_stimulus_from_test(self):
        """Delete only the selected stimulus from the current test's working order."""
        if not self.current_test_name:
            return
        selected_items = self.image_list.selectedItems()
        if not selected_items:
            return
        selected_row = self.image_list.row(selected_items[0])
        
        # Ensure working order exists
        if self.current_test_name not in self.working_orders:
            self.working_orders[self.current_test_name] = self.original_assets[self.current_test_name].copy()
        
        # Remove from working order
        if 0 <= selected_row < len(self.working_orders[self.current_test_name]):
            del self.working_orders[self.current_test_name][selected_row]
        
        self.update_image_list()
        self.update_apply_button_state()

    def get_custom_orders(self):
        """Return the current custom orders."""
        return self.custom_orders.copy()

    def select_test(self, test_name):
        """Select the given test in the test selector combo box."""
        index = self.test_selector.findText(test_name)
        if index != -1:
            self.test_selector.setCurrentIndex(index)

    def import_order_from_csv(self):
        """Import stimulus order from a CSV or Excel file."""
        if not self.current_test_name:
            return

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Import Order from CSV/Excel",
            "",
            "CSV/Excel Files (*.csv *.xlsx);;All Files (*)",
            options=options
        )

        if not file_name:
            return

        imported_order_names = []
        try:
            if file_name.lower().endswith('.csv'):
                import csv
                with open(file_name, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    for row in reader:
                        if row and len(row) > 0:
                            imported_order_names.append(row[0].strip())
            elif file_name.lower().endswith('.xlsx'):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    QMessageBox.critical(
                        self,
                        "Import Failed",
                        "openpyxl is required for Excel file import. Please install it with 'pip install openpyxl'.",
                        QMessageBox.Ok
                    )
                    return
                wb = load_workbook(file_name, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                    if row and row[0]:
                        imported_order_names.append(str(row[0]).strip())
            else:
                QMessageBox.critical(
                    self,
                    "Import Failed",
                    "Unsupported file type. Please select a .csv or .xlsx file.",
                    QMessageBox.Ok
                )
                return

            # Validate and build the order
            imported_order = []
            missing = []
            for name in imported_order_names:
                norm_name = self.normalize_name(name)
                img_obj = None
                if norm_name == self.normalize_name("craving_rating"):
                    img_obj = CravingRatingAsset()
                else:
                    for img in self.all_asset_objs:
                        if hasattr(img, 'filename'):
                            base = os.path.basename(img.filename)
                            base_no_ext = os.path.splitext(base)[0]
                            norm_base = self.normalize_name(base)
                            norm_base_no_ext = self.normalize_name(base_no_ext)
                            # Accept match if normalized names match
                            if norm_name == norm_base or norm_name == norm_base_no_ext:
                                img_obj = img
                                break
                if img_obj:
                    imported_order.append(img_obj)
                else:
                    missing.append(name)

            if missing:
                available_names = [os.path.basename(img.filename) for img in self.all_asset_objs if hasattr(img, 'filename')]
                QMessageBox.critical(
                    self,
                    "Import Failed",
                    "The following images were not found in the available assets:\n\n"
                    + "\n".join(missing)
                    + "\n\nAvailable assets for this test:\n"
                    + "\n".join(available_names),
                    QMessageBox.Ok
                )
                return

            if self.is_passive_test():
                imported_order = [img for img in imported_order if not isinstance(img, CravingRatingAsset)]
                original_craving = CravingRatingAsset()
                original_craving.is_original = True
                imported_order.append(original_craving)
            # Update the working order for the current test
            self.working_orders[self.current_test_name] = imported_order

            if not self.validate_passive_unique_stimulus_limit():
                return

            self.update_image_list()
            self.update_apply_button_state()  # Update button state to show changes can be applied

            QMessageBox.information(
                self,
                "Import Successful",
                f"Stimulus order imported successfully from '{os.path.basename(file_name)}'.",
                QMessageBox.Ok
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Import Failed",
                f"Failed to import stimulus order:\n{e}",
                QMessageBox.Ok
            )

    def normalize_name(self, name):
        """Normalize the asset name for consistent matching."""
        # Lowercase, remove any common image extension (case-insensitive), strip spaces and underscores
        name = name.strip().lower()
        name = re.sub(r'\.(jpg|jpeg|png|bmp|gif|tiff|webp)$', '', name, flags=re.IGNORECASE)
        name = name.replace('_', '').replace(' ', '')
        return name

    def is_current_order_applied(self):
        """Return True if the current working order matches the applied custom order."""
        if not self.current_test_name:
            return True  # No test selected, consider as "applied"
        
        # Ensure working order is synced with UI
        self.sync_working_order_with_ui()
        
        # Get working order
        working_order = self.working_orders.get(self.current_test_name, [])
        
        # Get applied order (custom order if exists, otherwise original order)
        if self.current_test_name in self.custom_orders:
            applied_order = self.custom_orders[self.current_test_name]
        else:
            applied_order = self.original_assets.get(self.current_test_name, [])
        
        return working_order == applied_order

    def update_apply_button_state(self):
        """Enable/disable the apply button based on whether the order is applied."""
        if hasattr(self, 'apply_button'):
            self.apply_button.setEnabled(not self.is_current_order_applied())

    def open_repetition_dialog(self):
        """Open dialog to set stimulus repetitions for images currently in the working order."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Set Stimulus Repetitions")
        layout = QFormLayout(dialog)

        # Get only images currently in the working order
        if self.current_test_name in self.working_orders:
            images = self.working_orders[self.current_test_name]
        elif self.current_test_name in self.custom_orders:
            images = self.custom_orders[self.current_test_name]
        else:
            images = self.original_assets.get(self.current_test_name, [])

        stimulus_names = []
        seen = set()
        for img in images:
            if hasattr(img, 'filename') and img.filename:
                base_name = os.path.splitext(os.path.basename(img.filename))[0]
                norm = self.normalize_name(base_name)
            elif hasattr(img, 'asset_type') and img.asset_type == "craving_rating":
                base_name = "craving_rating"
                norm = "craving_rating"
            else:
                base_name = "Image"
                norm = str(img)
            if norm not in seen:
                seen.add(norm)
                stimulus_names.append(base_name)

        edits = {}
        for name in stimulus_names:
            edit = QLineEdit()
            edit.setPlaceholderText("Repetitions (default 1)")
            layout.addRow(name, edit)
            edits[name] = edit

        # Add Set All button
        set_all_row = QHBoxLayout()
        set_all_edit = QLineEdit()
        set_all_edit.setPlaceholderText("Set all to...")
        set_all_button = QPushButton("Set All")
        set_all_button.clicked.connect(lambda: [
            edit.setText(set_all_edit.text()) for edit in edits.values() if set_all_edit.text().isdigit()
        ])
        set_all_row.addWidget(set_all_edit)
        set_all_row.addWidget(set_all_button)
        layout.addRow("Set all repetitions:", set_all_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        result = dialog.exec_()
        if result == QDialog.Accepted:
            self.stimulus_repetitions = {
                self.normalize_name(name): int(edits[name].text()) if edits[name].text().isdigit() else 1
                for name in stimulus_names
            }
            return True
        return False

    def get_stimulus_names_from_folders(self, alcohol_folder, non_alcohol_folder):
        """Get stimulus names from the specified folders."""
        supported_exts = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp')
        names = set()
        for folder in [alcohol_folder, non_alcohol_folder]:
            if folder and os.path.isdir(folder):
                for fname in os.listdir(folder):
                    if fname.lower().endswith(supported_exts):
                        base_name = os.path.splitext(fname)[0]
                        names.add(base_name)
        # Always include defaults if folders are empty
        if not names:
            names.update(["Beer", "Stella"])
        return sorted(names)
    
    def get_randomization_settings(self):
        """Get current randomization settings."""
        if hasattr(self, 'randomize_checkbox') and hasattr(self, 'seed_input'):
            randomize_cues = self.randomize_checkbox.isChecked()
            seed_text = self.seed_input.text().strip()
            seed = int(seed_text) if seed_text.isdigit() else seed_text if seed_text else None
            return randomize_cues, seed
        return False, None
    
    def get_repetitions_settings(self):
        """Get current repetitions settings."""
        if hasattr(self, 'repetition_checkbox') and self.repetition_checkbox.isChecked():
            return self.stimulus_repetitions
        return None

    def on_randomize_now_clicked(self):
        """Randomize and show the new order, including Craving Rating and repetitions."""
        if not self.randomize_checkbox.isChecked():
            QMessageBox.critical(
                self,
                "Randomization Disabled",
                "Please check the 'Randomize Alcohol/Non-Alcohol Cues' box before randomizing.",
                QMessageBox.Ok
            )
            return

        # Only show repetition dialog if the checkbox is checked
        if self.repetition_checkbox.isChecked():
            accepted = self.open_repetition_dialog()
            if not accepted:
                # User cancelled, do NOT randomize, keep working order unchanged
                return
            repetitions = self.stimulus_repetitions
        else:
            repetitions = None

        randomize_cues, seed = self.get_randomization_settings()
        if not self.current_test_name:
            return

        # Get images from current working order
        if self.current_test_name in self.working_orders:
            images = self.working_orders[self.current_test_name][:]
        else:
            if self.current_test_name in self.custom_orders:
                images = self.custom_orders[self.current_test_name][:]
            else:
                images = self.original_assets.get(self.current_test_name, [])[:]

        # Build a mapping from normalized name to image object (first found)
        name_to_img = {}
        for img in images:
            if hasattr(img, 'filename') and img.filename:
                base_name = os.path.splitext(os.path.basename(img.filename))[0]
                norm_name = self.normalize_name(base_name)
            elif hasattr(img, 'asset_type') and img.asset_type == "craving_rating":
                norm_name = self.normalize_name("craving_rating")  # <-- FIXED
            else:
                norm_name = str(img)
            if norm_name not in name_to_img:
                name_to_img[norm_name] = img

        # Use default repetitions if not specified
        if repetitions is None:
            # Default: one of each unique asset
            repeated_images = list(name_to_img.values())
        else:
            repeated_images = []
            for norm_name, count in repetitions.items():
                img = name_to_img.get(norm_name)
                if img:
                    repeated_images.extend([img] * count)

        # Shuffle the repeated images
        import random
        if seed is not None:
            random.seed(seed)
        random.shuffle(repeated_images)

        if self.is_passive_test():
            # Remove any existing craving asset, then append one
            repeated_images = [img for img in repeated_images if not isinstance(img, CravingRatingAsset)]
            original_craving = CravingRatingAsset()
            original_craving.is_original = True
            repeated_images.append(original_craving)

        # Update working order with randomized images
        self.working_orders[self.current_test_name] = repeated_images

        if not self.validate_passive_unique_stimulus_limit():
            return

        # Update the UI
        self.update_image_list()
        self.update_apply_button_state()

        QMessageBox.information(
            self,
            "Randomized!",
            "Stimulus order has been randomized in the working order. Click 'Apply Custom Order' to save this order if desired.",
            QMessageBox.Ok
        )
    def is_passive_test(self):
        return self.current_test_name and not self.current_test_name.lower().startswith("stroop")

    def validate_passive_unique_stimulus_limit(self):
        """Enforce passive viewing limit of max 8 unique non-craving stimuli."""
        if not self.is_passive_test() or not self.current_test_name:
            return True

        if self.current_test_name in self.working_orders:
            images = self.working_orders[self.current_test_name]
        elif self.current_test_name in self.custom_orders:
            images = self.custom_orders[self.current_test_name]
        else:
            images = self.original_assets.get(self.current_test_name, [])

        unique_names = set()
        for img in images:
            if isinstance(img, CravingRatingAsset):
                continue
            if hasattr(img, 'filename') and img.filename:
                base_name = os.path.splitext(os.path.basename(img.filename))[0]
                unique_names.add(self.normalize_name(base_name))
            elif hasattr(img, 'display_name') and img.display_name:
                unique_names.add(self.normalize_name(str(img.display_name)))
            else:
                unique_names.add(self.normalize_name(str(img)))

        if len(unique_names) > 8:
            QMessageBox.critical(
                self,
                "Too Many Unique Stimuli",
                f"Passive viewing tests support at most 8 unique stimuli. "
                f"Current order has {len(unique_names)} unique stimuli.\n\n"
                "Remove or merge stimuli until there are 8 or fewer unique entries.",
                QMessageBox.Ok
            )
            return False
        return True

    def goto_selected_test(self):
        """Ask for confirmation, then go to the selected test if confirmed."""
        if not self.current_test_name:
            QMessageBox.warning(
                self,
                "No Test Selected",
                "Please select a test from the dropdown first.",
                QMessageBox.Ok
            )
            return

            self.sync_working_order_with_ui()
            if not self.validate_passive_unique_stimulus_limit():
                return

        reply = QMessageBox.question(
            self,
            "Continue to Test",
            f"Are you sure you want to continue to the test:\n{self.current_test_name}\n"
            "Any unsaved changes to the stimulus order will be lost.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            # Call a method on the parent to show the selected test, if available
            if hasattr(self.parent, "show_test_frame"):
                self.parent.show_test_frame(self.current_test_name)
            else:
                QMessageBox.information(
                    self,
                    "Go to Test",
                    f"Would go to test: {self.current_test_name}\n(Implement show_test_frame in parent to enable navigation.)",
                    QMessageBox.Ok
                )
    def open_scent_assignment_dialog(self):
        if not self.current_test_name:
            return
        images = self.working_orders.get(self.current_test_name, [])
        dialog = QDialog(self)
        dialog.setWindowTitle("Assign Scent Numbers")
        layout = QFormLayout(dialog)
        scent_selectors = {}
        
        # Collect unique assets by filename to avoid showing duplicates
        seen_keys = set()
        unique_images = []
        for img in images:
            # Skip craving rating assets
            if isinstance(img, CravingRatingAsset):
                continue
            key = getattr(img, "filename", None)
            # Only add to unique_images if we haven't seen this filename before
            if key not in seen_keys:
                seen_keys.add(key)
                unique_images.append(img)
        
        # Create combo boxes only for unique assets
        for img in unique_images:
            display_name = getattr(img, "display_name", os.path.splitext(os.path.basename(getattr(img, "filename", "Image")))[0])
            combo = QComboBox()
            combo.addItem("None")
            for i in range(1, 9):
                combo.addItem(str(i))
            # Set current value
            key = getattr(img, "filename", None)
            scent = self.scent_numbers.get(key, None)
            if scent:
                combo.setCurrentText(str(scent))
            layout.addRow(display_name, combo)
            scent_selectors[key] = combo
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec_() == QDialog.Accepted:
            for key, combo in scent_selectors.items():
                val = combo.currentText()
                self.scent_numbers[key] = int(val) if val.isdigit() else None
            self.update_image_list()
    def validate_olfactory_scent_assignments(self, test_name):
        """
        Validate that all olfactory assets in the test have scent numbers assigned.
        Returns (is_valid, missing_scents_list)
        """
        if "olfactory" not in test_name.lower():
            # Not an olfactory test, so no validation needed
            return True, []
        
        images = self.working_orders.get(test_name, [])
        missing_scents = []
        
        for img in images:
            # Skip craving rating assets
            if isinstance(img, CravingRatingAsset):
                continue
            
            # Get the key for scent lookup
            key = getattr(img, "filename", None)
            
            # Check if this asset has a scent assigned
            scent = self.scent_numbers.get(key, None)
            if scent is None:
                display_name = getattr(img, "display_name", os.path.splitext(os.path.basename(key if key else "Unknown"))[0])
                missing_scents.append(display_name)
        
        return len(missing_scents) == 0, missing_scents
class CravingRatingAsset:
    def __init__(self):
        self.asset_type = "craving_rating"
        self.display_name = "craving_rating"